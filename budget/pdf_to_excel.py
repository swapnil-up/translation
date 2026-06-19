import argparse
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pdf2image import convert_from_path
from paddleocr import PaddleOCR

DEVANAGARI_RE = re.compile(r"[\u0900-\u097F]")
DIGIT_MAP = str.maketrans("०१२३४५६७८९", "0123456789")

TOTAL_KEYWORDS = {"जम्मा", "कुल", "जोड", "कूल"}
HEADER_KEYWORDS = {
    "अनुदान", "संकेत", "संख्या", "पृष्ठ", "निकायको", "विवरण",
    "चालु", "पूंजीगत", "वित्तीय", "व्यवस्था", "स्रोत", "जम्मा",
    "शीर्षक", "उपशीर्षक", "खर्च", "व्यय", "अनुमान", "सारांश",
    "वैदेशिक", "आन्तरिक", "ऋण", "ब्याज", "भुक्तानी",
    "नेपाल", "सरकार", "बजेट", "एकिकृत", "निकासा", "प्राथमिक",
    "रणनीति", "लैंगिक",
}

NOISE_CHARS = set("—−–-✓qoxmστSαT√")


def pdf_to_images(pdf_path: str, dpi: int = 300) -> list:
    print(f"[pdf]  {Path(pdf_path).name}  {dpi} dpi ...", file=sys.stderr, flush=True)
    imgs = convert_from_path(pdf_path, dpi=dpi)
    print(f"[pdf]  {len(imgs)} page(s)", file=sys.stderr, flush=True)
    return imgs


def extract_lines(images: list) -> list:
    ocr = PaddleOCR(
        text_recognition_model_name="devanagari_PP-OCRv5_mobile_rec",
        text_detection_model_name="PP-OCRv5_mobile_det",
        use_doc_orientation_classify=False,
        use_doc_unwarping=False,
        use_textline_orientation=False,
    )
    lines = []
    for page_idx, img in enumerate(images):
        print(f"[ocr]  Page {page_idx + 1}/{len(images)} ...", file=sys.stderr, flush=True)
        before = len(lines)
        arr = np.array(img)
        result = ocr.predict(arr)
        for record in result:
            texts = record["rec_texts"]
            scores = record["rec_scores"]
            boxes = record["rec_boxes"]
            for i in range(len(texts)):
                t = texts[i].strip()
                s = scores[i]
                if s < 0.3 or not t:
                    continue
                dev_ratio = len(DEVANAGARI_RE.findall(t)) / max(len(t), 1)
                if dev_ratio < 0.4 and len(t) >= 2 and s < 0.7:
                    continue
                x1, y1, x2, y2 = [int(boxes[i][j]) for j in range(4)]
                lines.append({
                    "y": y1, "x": x1, "w": max(x2 - x1, 60), "h": max(y2 - y1, 14),
                    "text": t, "page": page_idx,
                })
        added = len(lines) - before
        print(f"[ocr]   -> {added} lines", file=sys.stderr, flush=True)
    print(f"[ocr]  Total: {len(lines)} lines", file=sys.stderr, flush=True)
    return lines


def devanagari_to_arabic(text: str) -> str:
    return text.translate(DIGIT_MAP)


def parse_number(text: str):
    text = devanagari_to_arabic(text)
    cleaned = re.sub(r"[^\d,.\-]", "", text)
    cleaned = cleaned.replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def is_noise(t: str) -> bool:
    t = t.strip()
    if len(t) <= 1 and t in NOISE_CHARS:
        return True
    if t in ("", " "):
        return True
    return False


def cluster_columns(lines, bandwidth=70, min_count=2):
    if not lines:
        return []
    centers = np.sort(np.array([
        l["x"] + l["w"] / 2 for l in lines
        if not is_noise(l["text"]) and l["w"] > 20
    ]))
    if len(centers) == 0:
        return []
    clusters = []
    current = [centers[0]]
    for c in centers[1:]:
        if c - current[-1] <= bandwidth:
            current.append(c)
        else:
            if len(current) >= min_count:
                clusters.append(round(float(np.median(current))))
            current = [c]
    if len(current) >= min_count:
        clusters.append(round(float(np.median(current))))
    return sorted(clusters)


def merge_close_columns(cols, min_gap=60):
    if len(cols) < 2:
        return cols
    merged = [cols[0]]
    for c in cols[1:]:
        if c - merged[-1] < min_gap:
            merged[-1] = round((merged[-1] + c) / 2)
        else:
            merged.append(c)
    return merged


def is_two_column_toc(lines, page_width):
    texts = [l["text"] for l in lines if not is_noise(l["text"])]
    grant_headers = sum(1 for t in texts if t.strip() == "अनुदान")
    code_headers = sum(1 for t in texts if t.strip() == "संकेत")
    page_headers = sum(1 for t in texts if "पृष्ठ" in t or "संख्या" in t)
    has_suchi = any("सूची" in t for t in texts)
    has_anudan_x2 = grant_headers >= 2 and code_headers >= 1
    return has_suchi or (has_anudan_x2 and page_headers >= 2)


def group_rows(lines, y_tolerance=12):
    if not lines:
        return []
    remaining = sorted(lines, key=lambda l: (l["y"], l["x"]))
    rows = [[remaining[0]]]
    for l in remaining[1:]:
        if abs(l["y"] - rows[-1][-1]["y"]) <= y_tolerance:
            rows[-1].append(l)
        else:
            rows.append([l])
    # Merge rows with very few lines into adjacent rows
    merged = [rows[0]]
    for r in rows[1:]:
        non_noise = [l for l in r if not is_noise(l["text"])]
        if len(non_noise) <= 1 and merged:
            merged[-1].extend(r)
        else:
            merged.append(r)
    return merged


def build_grid(rows, column_centers, max_distance=150):
    if not column_centers:
        return [[" ".join(l["text"] for l in r if not is_noise(l["text"]))] for r in rows]

    def col_for_line(l):
        cx = l["x"] + l["w"] / 2
        distances = [abs(cx - cc) for cc in column_centers]
        best = min(range(len(distances)), key=lambda i: distances[i])
        if distances[best] <= max_distance:
            return best
        return None

    grid = []
    for row_lines in rows:
        row = [""] * len(column_centers)
        for l in row_lines:
            if is_noise(l["text"]):
                continue
            ci = col_for_line(l)
            if ci is not None:
                if row[ci]:
                    row[ci] += " "
                row[ci] += l["text"]
        grid.append(row)

    # Trim trailing empty columns
    if grid:
        max_col = 0
        for row in grid:
            for ci in range(len(row) - 1, -1, -1):
                if row[ci].strip():
                    max_col = max(max_col, ci + 1)
                    break
        grid = [row[:max_col] for row in grid]
    return grid


def is_header_row(row_texts):
    combined = " ".join(t for t in row_texts if t.strip()).strip()
    if not combined or len(combined) < 4:
        return False
    words = set(combined.split())
    hits = words & HEADER_KEYWORDS
    return len(hits) >= 1


def is_total_row(row_texts):
    combined = " ".join(t for t in row_texts if t.strip()).strip()
    return any(kw in combined for kw in TOTAL_KEYWORDS)


def process_page(lines, page_idx, page_width):
    filtered = [l for l in lines if not is_noise(l["text"])]
    if not filtered or len(filtered) < 3:
        return {"type": "empty"}

    # TOC pages: two-column list of ministries
    if is_two_column_toc(filtered, page_width):
        return process_two_column(filtered, page_width)

    columns = cluster_columns(filtered)
    columns = merge_close_columns(columns)

    x_span = max(l["x"] for l in filtered) - min(l["x"] for l in filtered) if filtered else 0
    if len(columns) < 2 or x_span < page_width * 0.25:
        texts = [l["text"] for l in sorted(filtered, key=lambda x: (x["y"], x["x"]))]
        return {"type": "text", "data": "\n".join(texts)}

    rows = group_rows(filtered)
    grid = build_grid(rows, columns)

    # Verify the grid has enough tabular structure
    data_rows = sum(1 for r in grid if sum(1 for c in r if c.strip()) >= 2)
    numeric_cols = sum(1 for ci in range(len(columns)) if any(parse_number(grid[ri][ci]) is not None for ri in range(len(grid))))
    if data_rows < 3:
        texts = [l["text"] for l in sorted(filtered, key=lambda x: (x["y"], x["x"]))]
        return {"type": "text", "data": "\n".join(texts)}
    if len(grid) <= len(columns) + 1 and numeric_cols < 3:
        texts = [l["text"] for l in sorted(filtered, key=lambda x: (x["y"], x["x"]))]
        return {"type": "text", "data": "\n".join(texts)}

    while grid and all(c == "" for c in grid[-1]):
        grid.pop()
    while grid and all(c == "" for c in grid[0]):
        grid.pop(0)

    header_count = 0
    for r in grid:
        if is_header_row(r):
            header_count += 1
        else:
            break

    return {
        "type": "table",
        "columns": columns,
        "grid": grid,
        "header_count": header_count,
    }


def process_two_column(lines, page_width):
    mid = page_width / 2
    left_lines = sorted(
        [l for l in lines if l["x"] + l["w"] / 2 < mid],
        key=lambda x: (x["y"], x["x"]),
    )
    right_lines = sorted(
        [l for l in lines if l["x"] + l["w"] / 2 >= mid],
        key=lambda x: (x["y"], x["x"]),
    )

    left_cols = merge_close_columns(cluster_columns(left_lines))
    right_cols = merge_close_columns(cluster_columns(right_lines))

    left_rows = group_rows(left_lines)
    right_rows = group_rows(right_lines)

    left_grid = build_grid(left_rows, left_cols)
    right_grid = build_grid(right_rows, right_cols)

    # Trim trailing empty columns
    for g in (left_grid, right_grid):
        while g and all(c == "" for c in g[-1]):
            g.pop()
    if left_grid:
        max_c = max(next((ci for ci in range(len(left_grid[0])-1, -1, -1) if any(r[ci].strip() for r in left_grid)), 0) + 1, 1)
        left_grid = [r[:max_c] for r in left_grid]
    if right_grid:
        max_c = max(next((ci for ci in range(len(right_grid[0])-1, -1, -1) if any(r[ci].strip() for r in right_grid)), 0) + 1, 1)
        right_grid = [r[:max_c] for r in right_grid]

    return {
        "type": "two_column_table",
        "left": {"columns": left_cols, "grid": left_grid},
        "right": {"columns": right_cols, "grid": right_grid},
    }


def cross_verify(grid, columns):
    if not grid or len(grid) < 4:
        return []
    if not columns:
        return []

    # Find numeric columns with at least 3 data values
    col_data = []
    for ci in range(len(columns)):
        vals = []
        for ri, row in enumerate(grid):
            if not is_total_row(row):
                p = parse_number(row[ci])
                if p is not None:
                    vals.append((ri, p))
        if len(vals) >= 3:
            col_data.append((ci, vals))

    total_rows = [ri for ri, row in enumerate(grid) if is_total_row(row)]

    checks = []
    for ci, vals in col_data:
        data_vals_above = {ri: v for ri, v in vals}
        for tr in total_rows:
            stated = parse_number(grid[tr][ci])
            if stated is None or abs(stated) < 1:
                continue
            above = [v for ri, v in vals if ri < tr]
            if len(above) < 2:
                continue
            computed = sum(above)
            diff = abs(computed - stated)
            rel_diff = diff / max(abs(stated), 1)
            # Only flag significant discrepancies (> 1% or > 50000)
            if diff > 50000 and rel_diff > 0.001:
                checks.append({
                    "row": tr,
                    "computed": computed,
                    "stated": stated,
                    "diff": computed - stated,
                })

    return checks


def write_grid(ws, grid, start_row=1, start_col=1, header_count=0):
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    alt_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ri, row in enumerate(grid):
        for ci, text in enumerate(row):
            cell = ws.cell(row=start_row + ri, column=start_col + ci, value=text)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

            if ri < header_count:
                cell.fill = header_fill
                cell.font = Font(bold=True, size=10)
            elif is_total_row(row):
                cell.fill = total_fill
                cell.font = Font(bold=True, size=10)
            elif ri % 2 == 1:
                cell.fill = alt_fill


def export_to_excel(pages_data, output_path):
    wb = Workbook()
    wb.remove(wb.active)

    for pd_ in pages_data:
        ptype = pd_["type"]
        pnum = pd_.get("page_num", 1)

        if ptype == "empty":
            continue

        if ptype == "text":
            ws = wb.create_sheet(title=f"Page {pnum}")
            ws.cell(row=1, column=1, value=pd_["data"])

        elif ptype == "table":
            ws = wb.create_sheet(title=f"Page {pnum}")
            write_grid(ws, pd_["grid"], header_count=pd_["header_count"])

            for ci in range(len(pd_["grid"][0]) if pd_["grid"] else 0):
                max_len = 0
                for ri in range(len(pd_["grid"])):
                    if ci < len(pd_["grid"][ri]):
                        val = str(pd_["grid"][ri][ci])
                        max_len = max(max_len, len(val))
                col_letter = chr(65 + ci) if ci < 26 else f"A{ci - 25}"
                ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

            cross = pd_.get("cross_checks", [])
            if cross:
                cr = len(pd_["grid"]) + 3
                ws.cell(row=cr, column=1, value="Cross-Verification Checks").font = Font(bold=True, color="CC0000")
                cr += 1
                for col_name in ["Row", "Computed", "Stated", "Diff"]:
                    cell = ws.cell(row=cr, column=1 + ["Row", "Computed", "Stated", "Diff"].index(col_name), value=col_name)
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                cr += 1
                for ch in cross:
                    ws.cell(row=cr, column=1, value=ch["row"] + 1)
                    ws.cell(row=cr, column=2, value=round(ch["computed"], 2))
                    ws.cell(row=cr, column=3, value=round(ch["stated"], 2))
                    ws.cell(row=cr, column=4, value=round(ch["diff"], 2))
                    cr += 1

        elif ptype == "two_column_table":
            ws = wb.create_sheet(title=f"Page {pnum}")
            left = pd_["left"]
            right = pd_["right"]

            write_grid(ws, left["grid"], start_col=1)

            if right["columns"] and right["grid"]:
                left_col_count = len(left["grid"][0]) if left["grid"] else 3
                offset = left_col_count + 3
                write_grid(ws, right["grid"], start_col=offset)

    # Ensure at least one visible sheet exists
    if not any(ws.sheet_state == "visible" for ws in wb.worksheets):
        ws = wb.create_sheet(title="Output")
        ws.cell(row=1, column=1, value="No structured tables found in the processed pages.")

    wb.save(output_path)
    print(f"[excel] Written to {output_path}", file=sys.stderr)


def export_to_csv(pages_data, out_dir, stem):
    for pd_ in pages_data:
        ptype = pd_["type"]
        pnum = pd_.get("page_num", 1)
        if ptype == "empty":
            continue

        def grid_to_csv_rows(grid):
            rows = []
            for row in grid:
                csv_cells = []
                for cell in row:
                    s = str(cell).replace('"', '""')
                    if "," in s or '"' in s or "\n" in s:
                        csv_cells.append(f'"{s}"')
                    else:
                        csv_cells.append(s)
                rows.append(",".join(csv_cells))
            return rows

        if ptype == "text":
            path = Path(out_dir) / f"{stem}-page{pnum}.csv"
            path.write_text(f'"{pd_["data"].replace(chr(34), chr(34)+chr(34))}"', encoding="utf-8")
            print(f"[csv]   {path.name}", file=sys.stderr)

        elif ptype == "table":
            path = Path(out_dir) / f"{stem}-page{pnum}.csv"
            rows = grid_to_csv_rows(pd_["grid"])
            cross = pd_.get("cross_checks", [])
            if cross:
                rows.append("")
                rows.append("Cross-Verification Checks")
                rows.append("Row,Computed,Stated,Diff")
                for ch in cross:
                    rows.append(f"{ch['row']+1},{round(ch['computed'],2)},{round(ch['stated'],2)},{round(ch['diff'],2)}")
            path.write_text("\n".join(rows), encoding="utf-8")
            print(f"[csv]   {path.name}  ({len(pd_['grid'])} rows)", file=sys.stderr)

        elif ptype == "two_column_table":
            left_rows = grid_to_csv_rows(pd_["left"]["grid"])
            right_rows = grid_to_csv_rows(pd_["right"]["grid"])
            max_l = max(len(r) for r in left_rows) if left_rows else 1
            # Interleave left and right: left cols, gap, right cols
            combined = []
            for i in range(max(len(left_rows), len(right_rows))):
                l = left_rows[i] if i < len(left_rows) else "," * max_l
                r = right_rows[i] if i < len(right_rows) else ""
                combined.append(l + ",," + r)
            path = Path(out_dir) / f"{stem}-page{pnum}.csv"
            path.write_text("\n".join(combined), encoding="utf-8")
            print(f"[csv]   {path.name}  ({len(left_rows)}L+{len(right_rows)}R rows)", file=sys.stderr)


def export_to_sqlite(pages_data, output_path):
    import sqlite3
    conn = sqlite3.connect(str(output_path))
    conn.execute("PRAGMA journal_mode=WAL")

    # Convert grid column to a safe column name
    def col_name(idx):
        return f"col_{idx:03d}"

    for pd_ in pages_data:
        ptype = pd_["type"]
        pnum = pd_.get("page_num", 1)
        if ptype == "empty":
            continue

        tbl = f"page_{pnum}"

        if ptype == "text":
            conn.execute(f"CREATE TABLE IF NOT EXISTS {tbl} (content TEXT)")
            conn.execute(f"INSERT INTO {tbl} (content) VALUES (?)", (pd_["data"],))

        elif ptype == "table":
            grid = pd_["grid"]
            ncols = len(grid[0]) if grid else 0
            col_names = [col_name(i) for i in range(ncols)]
            cols_sql = ", ".join(col_names)
            ph = ", ".join("?" for _ in range(4 + ncols))
            conn.execute(f"CREATE TABLE IF NOT EXISTS {tbl} (row_id INTEGER, page_type TEXT, is_header INTEGER, is_total INTEGER, {cols_sql})")
            for ri, row in enumerate(grid):
                vals = [row[ci] if ci < len(row) else "" for ci in range(ncols)]
                conn.execute(
                    f"INSERT INTO {tbl} VALUES ({ph})",
                    [ri, "table", 1 if ri < pd_.get("header_count", 0) else 0, 1 if is_total_row(row) else 0] + vals,
                )
            cross = pd_.get("cross_checks", [])
            if cross:
                conn.execute(f"CREATE TABLE IF NOT EXISTS cross_verify_page_{pnum} (row INTEGER, computed REAL, stated REAL, diff REAL)")
                for ch in cross:
                    conn.execute(f"INSERT INTO cross_verify_page_{pnum} VALUES (?, ?, ?, ?)",
                                 (ch["row"] + 1, ch["computed"], ch["stated"], ch["diff"]))

        elif ptype == "two_column_table":
            for side, side_label in [(pd_["left"], "L"), (pd_["right"], "R")]:
                grid = side["grid"]
                ncols = len(grid[0]) if grid else 0
                col_names = [f"{side_label}_{col_name(i)}" for i in range(ncols)]
                tbl_side = f"{tbl}_{side_label}"
                cols_sql = ", ".join(col_names)
                conn.execute(f"CREATE TABLE IF NOT EXISTS {tbl_side} (row_id INTEGER, {cols_sql})")
                for ri, row in enumerate(grid):
                    conn.execute(
                        f"INSERT INTO {tbl_side} (row_id, {cols_sql}) VALUES (?, {', '.join('?' for _ in range(ncols))})",
                        [ri] + [row[ci] if ci < len(row) else "" for ci in range(ncols)],
                    )

    conn.commit()
    conn.close()
    print(f"[sqlite] Written to {output_path}", file=sys.stderr)


def main():
    parser = argparse.ArgumentParser(
        description="Convert Nepali government budget PDF to table formats (Excel / CSV / SQLite)"
    )
    parser.add_argument("pdf", help="Path to input PDF file")
    parser.add_argument("--dpi", type=int, default=300, help="DPI for PDF rendering")
    parser.add_argument("--output", "-o", help="Output path (default: output/<pdf-name>.*)")
    parser.add_argument("--max-pages", type=int, default=None, help="Process only first N pages")
    fmt = parser.add_mutually_exclusive_group()
    fmt.add_argument("--xlsx", action="store_true", help="Export to Excel (default)")
    fmt.add_argument("--csv", action="store_true", help="Export to CSV files (one per page)")
    fmt.add_argument("--sqlite", action="store_true", help="Export to SQLite database")
    args = parser.parse_args()

    pdf = Path(args.pdf)
    if not pdf.exists():
        print(f"Error: {pdf} not found", file=sys.stderr)
        sys.exit(1)

    imgs = pdf_to_images(str(pdf), dpi=args.dpi)
    if args.max_pages:
        imgs = imgs[: args.max_pages]
        print(f"[stage] Limited to {args.max_pages} pages", file=sys.stderr)

    lines = extract_lines(imgs)

    pages = defaultdict(list)
    for l in lines:
        pages[l["page"]].append(l)

    pages_data = []
    for page_idx in sorted(pages.keys()):
        page_lines = pages[page_idx]
        page_width = imgs[page_idx].width if page_idx < len(imgs) else 3508

        result = process_page(page_lines, page_idx, page_width)
        result["page_num"] = page_idx + 1

        if result["type"] == "table":
            checks = cross_verify(result["grid"], result["columns"])
            if checks:
                result["cross_checks"] = checks
                print(f"[verify] Page {page_idx + 1}: {len(checks)} discrepancy(ies)", file=sys.stderr)
                for ch in checks:
                    print(f"  Row {ch['row']+1}: computed={ch['computed']:.0f} stated={ch['stated']:.0f} diff={ch['diff']:.0f}", file=sys.stderr)
        elif result["type"] == "two_column_table":
            print(f"[page]  Page {page_idx + 1}: two-column TOC ({len(result['left']['grid'])}L / {len(result['right']['grid'])}R rows)", file=sys.stderr)
        elif result["type"] == "text":
            print(f"[page]  Page {page_idx + 1}: text ({len(result['data'].split(chr(10)))} lines)", file=sys.stderr)
        elif result["type"] == "empty":
            print(f"[page]  Page {page_idx + 1}: empty", file=sys.stderr)
        else:
            print(f"[page]  Page {page_idx + 1}: {result['type']}", file=sys.stderr)

        if result["type"] == "table":
            print(f"[page]  -> {len(result['columns'])} cols, {len(result['grid'])} rows, {result['header_count']} header(s)", file=sys.stderr)

        pages_data.append(result)

    stem = pdf.stem
    Path("output").mkdir(exist_ok=True)

    use_csv = args.csv
    use_sqlite = args.sqlite

    if use_csv:
        export_to_csv(pages_data, "output", stem)
    elif use_sqlite:
        output_path = args.output or str(Path("output") / f"{stem}.db")
        export_to_sqlite(pages_data, output_path)
    else:
        output_path = args.output or str(Path("output") / f"{stem}.xlsx")
        export_to_excel(pages_data, output_path)


if __name__ == "__main__":
    main()
