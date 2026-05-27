#!/usr/bin/env python
"""
pdf_to_excel_v2.py — Vector-layer Nepali budget PDF extraction pipeline.

Uses pdfplumber's built-in table detection for robust column/row parsing,
with CID→Unicode font mapping and Preeti→Unicode decoding.

Architecture:
  pdfplumber tables → font-aware cell decoding → row classification →
  hierarchy flattening → validation → XLSX / CSV / SQLite export
"""

import argparse
import os
import re
import sqlite3
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Preeti → Unicode mapping ────────────────────────────────────────────────
PREETI_MAP = {
    ' ': ' ', '!': '!', '"': '"', '#': '॥', '$': '₹', '%': '%', '&': '&',
    "'": "'", '(': '(', ')': ')', '*': '*', '+': '+', ',': ',', '-': '-',
    '.': '.', '/': '/',
    '0': '०', '1': '१', '2': '२', '3': '३', '4': '४',
    '5': '५', '6': '६', '7': '७', '8': '८', '9': '९',
    ':': ':', ';': ';', '<': '<', '=': '=', '>': '>', '?': '?',
    '@': '@',
    'A': 'अ', 'B': 'ब', 'C': 'क', 'D': 'ड', 'E': 'ए',
    'F': 'फ', 'G': 'ग', 'H': 'ह', 'I': 'इ', 'J': 'ज',
    'K': 'ख', 'L': 'ल', 'M': 'म', 'N': 'न', 'O': 'ओ',
    'P': 'प', 'Q': 'क', 'R': 'र', 'S': 'स', 'T': 'त',
    'U': 'उ', 'V': 'व', 'W': 'व', 'X': 'क्ष', 'Y': 'य', 'Z': 'ज्ञ',
    '[': '[', '\\': '\\', ']': ']', '^': '^', '_': '_',
    '`': '`',
    'a': 'अ', 'b': 'ब', 'c': 'क', 'd': 'ड', 'e': 'ए',
    'f': 'फ', 'g': 'ग', 'h': 'ह', 'i': 'इ', 'j': 'ज',
    'k': 'क', 'l': 'ल', 'm': 'म', 'n': 'न', 'o': 'ओ',
    'p': 'प', 'q': 'क', 'r': 'र', 's': 'स', 't': 'त',
    'u': 'उ', 'v': 'व', 'w': 'व', 'x': 'क्ष', 'y': 'य', 'z': 'ज्ञ',
    '{': '{', '|': '|', '}': '}', '~': '~',
}

# ── Constants ────────────────────────────────────────────────────────────────
DEVANAGARI_RE = re.compile(r"[\u0900-\u097F]")
CID_RE = re.compile(r'^\(cid:(\d+)\)$')
DIGIT_MAP = str.maketrans("०१२३४५६७८९", "0123456789")
TOTAL_KEYWORDS = {"जम्मा", "कुल", "जोड", "कूल", "योग"}
GRAND_TOTAL_KEYWORDS = {"कुल जम्मा", "कूल जम्मा", "जम्मा जम्मा"}

# Scale detection: if the page header says "रु हजारमा" multiply everything by 1,000
# "रु लाखमा" → 100,000
SCALE_PATTERNS = [
    (re.compile(r'हजार', re.UNICODE), 1000),
    (re.compile(r'लाख', re.UNICODE), 100_000),
    (re.compile(r'करोड', re.UNICODE), 10_000_000),
]

# TOC / non-budget page keywords
TOC_KEYWORDS = {"सूची-पत्र", "अनुक्रमणिका", "विषय सूची", "सूचीपत्र"}

# Known broken Devanagari words from legacy font mapping.
# These are observed mis-rendered words, not theoretical conjunct sequences.
WORD_FIXES = {
    # "प्र" rendered as "पर्" or "पर्द"
    'पर्देश': 'प्रदेश',
    'पर्मखु': 'प्रमुख',
    'पराकृतक': 'प्राकृतिक',
    'परशासन': 'प्रशासन',
    'परकाशन': 'प्रकाशन',
    'परविधान': 'प्रविधान',
    'परतिनिधि': 'प्रतिनिधि',
    'परणाली': 'प्रणाली',
    'परवन्ध': 'प्रबन्ध',
    'रापत': 'राष्ट्रपति',
    'उपरापत': 'उपराष्ट्रपति',
    'कायार्लय': 'कार्यालय',
    'कायार्क्रम': 'कार्यक्रम',
    'अनसन्धान': 'अनुसन्धान',
    'अनसु न्धान': 'अनुसन्धान',
    'दरुपयोग': 'दुरुपयोग',
    'सरोत': 'स्रोत',
    'संिचत': 'सञ्चित',
    'अिख्तयार': 'अख्तियार',
    'परक्षक': 'परीक्षक',
    'बैदेिशक': 'वैदेशिक',
    'बैदिशक': 'वैदेशिक',
    'अनदान': 'अनुदान',
    'अनदु ान': 'अनुदान',
    'ववध': 'विविध',
    'ववरण': 'विवरण',
    'नवाचन': 'निर्वाचन',
    'नवाचर्न': 'निर्वाचन',
    'महला': 'महिला',
    'परषद्': 'परिषद्',
    'आदवासी': 'आदिवासी',
    'जनजात': 'जनजाति',
    'मिुस्लम': 'मुस्लिम',
    'वनयोजन': 'विनियोजन',
    'पंजीगत': 'पूँजीगत',
    'पंजू ीगत': 'पूँजीगत',
    'आथकर्': 'आर्थिक',
    'अथकर्': 'आर्थिक',
    'साराशं': 'सारांश',
    'बहपु क्षीय': 'बहुपक्षीय',
    'बहपक्षीय': 'बहुपक्षीय',
    'स्पक्षीय': 'द्विपक्षीय',
    'सशोधन': 'संशोधन',
    'सशोधत': 'संशोधित',
    'संशोधत': 'संशोधित',
    'यथाथ': 'यथार्थ',
    'खचर्': 'खर्च',
    'खच र्': 'खर्च',
    'अनमान': 'अनुमान',
    'अनमु ान': 'अनुमान',
    'वषर्': 'वर्ष',
    'भु ानी': 'भुक्तानी',
    'आन्तरक': 'आन्तरिक',
    'मन्तरालय': 'मन्त्रालय',
    'मन्तर्ालय': 'मन्त्रालय',
    'अधकार': 'अधिकार',
    'सरोत': 'स्रोत',
    'सर्ोत': 'स्रोत',
    'दरुपयोग': 'दुरुपयोग',
    'दरुु पयोग': 'दुरुपयोग',
    'वैदेिशक': 'वैदेशिक',
    '(पक्षीय)': '(द्विपक्षीय)',
    'राय ': 'राष्ट्रिय ',
    'थ र् म': 'र्थ म',
    'थ र्-': 'र्थ-',
    'थ र् ': 'र्थ ',
    'थ र्': 'र्थ',
    'स्यात': 'स्वास्थ्य',
}


# ═══════════════════════════════════════════════════════════════════════════════
#  Font Mapper
# ═══════════════════════════════════════════════════════════════════════════════

class FontMapper:
    def __init__(self):
        self.cid_to_uni: dict[int, str] = {}
        self._load_fallback()

    def _load_fallback(self):
        self.cid_to_uni.update({
            94: 'इ',       107: 'औ',      134: 'ाल',
            163: 'ज्ञ',     207: 'ज्',     216: 'थ्य',
            217: '्द',      222: 'व्य',     227: 'ल्या',
            274: 'ग्री',    295: 'भ्र',     302: 'श्र',
            429: 'द्यो',    436: 'ष्ट',     437: 'ष्',
            461: 'ङ्गि',    548: 'न',       555: 'भौ',
            561: 'ँ',       680: 'ड्ड',     694: 'ञ्च',
        })

    def parse_cmap(self, text: str) -> dict[int, str]:
        mapping = {}
        in_bfchar = False
        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue
            m = re.match(r'^(\d+)\s+beginbfchar$', line)
            if m:
                in_bfchar = True
                continue
            if line == 'endbfchar':
                in_bfchar = False
                continue
            if in_bfchar:
                parts = line.split()
                if len(parts) >= 2:
                    cid_hex = parts[0].strip('<>')
                    uni_hex = parts[1].strip('<>')
                    try:
                        cid = int(cid_hex, 16)
                    except ValueError:
                        continue
                    if len(uni_hex) <= 4:
                        try:
                            cp = int(uni_hex, 16)
                            if cp >= 0x20:
                                mapping[cid] = chr(cp)
                        except ValueError:
                            pass
                    else:
                        chars = []
                        for off in range(0, len(uni_hex), 4):
                            try:
                                cp = int(uni_hex[off:off+4], 16)
                                if cp >= 0x20:
                                    chars.append(chr(cp))
                            except ValueError:
                                pass
                        if chars:
                            mapping[cid] = ''.join(chars)
        return mapping

    def ingest_cmap(self, text: str):
        self.cid_to_uni.update(self.parse_cmap(text))

    def decode_text(self, text: str, fontname: str = '') -> str:
        if not text:
            return text
        lines = text.split('\n')
        decoded = []
        for line in lines:
            parts = []
            for token in re.findall(r'\(cid:\d+\)|.', line):
                m = CID_RE.match(token)
                if m:
                    cid = int(m.group(1))
                    parts.append(self.cid_to_uni.get(cid, f'[CID:{cid}]'))
                elif 'Preeti' in fontname:
                    parts.append(PREETI_MAP.get(token, token))
                else:
                    parts.append(token)
            decoded.append(''.join(parts))
        return '\n'.join(decoded)


# ═══════════════════════════════════════════════════════════════════════════════
#  Text Sanitizers
# ═══════════════════════════════════════════════════════════════════════════════

def sanitize_devanagari(text: str) -> str:
    """Post-process decoded Devanagari text to fix broken words
    caused by legacy font glyph decomposition."""
    text = text.replace('\n', ' ').replace('\ufffd', '')
    # Apply known word fixes (longest-first to avoid partial matches)
    for broken, correct in sorted(WORD_FIXES.items(),
                                  key=lambda x: -len(x[0])):
        text = text.replace(broken, correct)
    return text.strip()


def detect_scale_multiplier(page_text: str) -> int:
    """Detect if page header specifies a unit scale (हजार, लाख, करोड).
    Returns the multiplier (1, 1000, 100000, 10_000_000)."""
    for pattern, multiplier in SCALE_PATTERNS:
        if pattern.search(page_text):
            return multiplier
    return 1


def is_valid_budget_page(page_text: str) -> bool:
    """Gatekeeper: skip TOC and non-budget pages."""
    for kw in TOC_KEYWORDS:
        if kw in page_text:
            return False
    # Must have at least one budget keyword
    if 'जम्मा' not in page_text and 'खर्च' not in page_text:
        return False
    return True


# ═══════════════════════════════════════════════════════════════════════════════
#  Extraction
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class PageTable:
    page_num: int
    bbox: tuple
    header: list[str]
    rows: list[list[str]]


# ═══════════════════════════════════════════════════════════════════════════════
#  Budget Classification (State Machine)
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class BudgetItem:
    page: int
    section: str
    code: str
    description: str
    year_actual: float | None = None
    year_revised: float | None = None
    year_estimate: float | None = None
    total: float | None = None
    current_exp: float | None = None
    capital_exp: float | None = None
    financial: float | None = None
    is_total: bool = False
    is_section_header: bool = False
    is_grand_total: bool = False


def parse_number(text: str, scale: int = 1) -> float | None:
    """Parse a Nepali-format number, optionally scaled by unit multiplier."""
    if not text:
        return None
    text = text.translate(DIGIT_MAP)
    cleaned = re.sub(r"[^\d,.\-]", "", text)
    cleaned = cleaned.replace(",", "")
    if not cleaned:
        return None
    try:
        return float(cleaned) * scale
    except ValueError:
        return None


def is_total_text(text: str) -> bool:
    return any(kw in text for kw in TOTAL_KEYWORDS)


def is_grand_total_text(text: str) -> bool:
    return any(kw in text for kw in GRAND_TOTAL_KEYWORDS)


def parse_long_code_row(row_text: str, scale: int = 1) -> dict | None:
    """Parse a row where pdfplumber merged all cells into one.
    Extracts budget code (first token), trailing numeric values,
    and description (everything between code and numbers).
    Caps at 6 numeric values to discard trailing page refs like 'P1 09'."""
    tokens = row_text.strip().split()
    if len(tokens) < 2:
        return None
    code = tokens[0].translate(DIGIT_MAP)
    if not re.match(r'^\d{5,9}$', code):
        return None

    # Find where numbers start (first purely-numeric token after code)
    num_start = None
    for i in range(1, len(tokens)):
        cand = tokens[i].translate(DIGIT_MAP)
        if re.match(r'^\d+$', cand):
            num_start = i
            break

    if num_start is None:
        # No numbers found — entire remainder is description
        return {
            'code': tokens[0],
            'description': ' '.join(tokens[1:]),
            'numeric_values': [],
        }

    description = ' '.join(tokens[1:num_start])
    # Parse all trailing numeric tokens, limited to 6 (year_actual..capital_exp)
    raw_numbers = tokens[num_start:]
    numeric_values = []
    for t in raw_numbers:
        v = parse_number(t, scale)
        if v is not None:
            numeric_values.append(v)
            if len(numeric_values) >= 6:
                break

    return {
        'code': tokens[0],
        'description': description,
        'numeric_values': numeric_values,
    }


def finalize_devanagari_spelling(text: str) -> str:
    """Post-process sanitized Devanagari for remaining cosmetic CID gaps."""
    if not text:
        return text
    cosmetic_fixes = {
        'हनु े': 'हुने',
        'चाल ु': 'चालु',
        'नवाचर् न': 'निर्वाचन',
        'नबाचन': 'निर्वाचन',
        'पर्ाकृतक': 'प्राकृतिक',
        'पर्ाकृतिक': 'प्राकृतिक',
        'दलत': 'दलित',
        'राष्ट्रिय दलत': 'राष्ट्रिय दलित',
    }
    for broken, correct in sorted(cosmetic_fixes.items(),
                                  key=lambda x: -len(x[0])):
        text = text.replace(broken, correct)
    # Fix trailing spurious spaces before vowel signs (after spell fixes)
    text = text.replace(' ु', 'ु').replace(' े', 'े').replace(' ी', 'ी')
    return text


def classify_budget_table(table: PageTable,
                          default_section: str = '',
                          scale: int = 1) -> list[BudgetItem]:
    """Apply state machine to classify rows in a budget table.
    Handles both detail tables (col 0=code, col 1=desc) and
    summary tables (col 0=section name, numeric values start at col 1).
    `default_section` is carried over from the previous page."""
    items: list[BudgetItem] = []
    current_section = default_section

    # Detect table type from first data row
    is_detail = False
    for row in table.rows:
        first = (row[0] or '').strip()
        if first:
            is_detail = bool(re.match(r'^\d+', first.translate(DIGIT_MAP)))
            break

    for ri, row in enumerate(table.rows):
        combined = ' '.join(c.strip() for c in row if c.strip()).strip()
        if not combined:
            continue

        first = row[0].strip() if row else ''
        second = row[1].strip() if len(row) > 1 else ''
        is_total = first and is_total_text(first)
        is_grand = first and is_grand_total_text(first)

        # Detect rows where pdfplumber merged cells (single cell, long code)
        long_code_match = re.match(r'^(\d{8,9})\b', first.translate(DIGIT_MAP))
        use_regex = is_detail and long_code_match and len(row) <= 2

        if use_regex:
            parsed = parse_long_code_row(combined, scale)
            if parsed:
                code = parsed['code']
                description = finalize_devanagari_spelling(
                    sanitize_devanagari(parsed['description']))
                numeric_values = parsed['numeric_values']
                is_section = False
                if '-' in description or '–' in description:
                    desc_section = description.split('-')[0].split('–')[0].strip()
                    if desc_section and len(desc_section) > 2:
                        current_section = desc_section
            else:
                use_regex = False

        if not use_regex:
            numeric_values = []
            if is_detail:
                code_match = re.match(r'^(\d+(?:\.\d+)?)',
                                      first.translate(DIGIT_MAP))
                code = code_match.group(1) if code_match else ''
                description = finalize_devanagari_spelling(
                    sanitize_devanagari(second if second else first))
                is_section = False
                if '-' in description or '–' in description:
                    desc_section = description.split('-')[0].split('–')[0].strip()
                    if desc_section and len(desc_section) > 2:
                        current_section = desc_section
                for ci in range(2, len(row)):
                    v = parse_number(row[ci], scale)
                    if v is not None:
                        numeric_values.append(v)
            else:
                code = ''
                if is_total:
                    description = finalize_devanagari_spelling(
                        sanitize_devanagari(combined))
                else:
                    description = finalize_devanagari_spelling(
                        sanitize_devanagari(first))
                    if first and first not in ('', ' '):
                        current_section = description
                is_section = not is_total and bool(first) and first not in ('', ' ')
                for ci in range(1, len(row)):
                    v = parse_number(row[ci], scale)
                    if v is not None:
                        numeric_values.append(v)

        item = BudgetItem(
            page=table.page_num + 1,
            section=current_section,
            code=code,
            description=description,
            is_total=is_total,
            is_section_header=is_section and not is_total,
            is_grand_total=is_grand,
        )

        n_num = len(numeric_values)
        if n_num >= 1:
            item.year_actual = numeric_values[0]
        if n_num >= 2:
            item.year_revised = numeric_values[1]
        if n_num >= 3:
            item.year_estimate = numeric_values[2]
        if n_num >= 4:
            item.total = numeric_values[3]
        if n_num >= 5:
            item.current_exp = numeric_values[4]
        if n_num >= 6:
            item.capital_exp = numeric_values[5]
        if n_num >= 7:
            item.financial = numeric_values[6]

        items.append(item)

    return items


# ═══════════════════════════════════════════════════════════════════════════════
#  Validation
# ═══════════════════════════════════════════════════════════════════════════════

def cross_verify(items: list[BudgetItem]) -> list[dict]:
    """Sum budget items per section and compare to stated totals.
    Page-level (grand) totals are verified against ALL page items."""
    checks = []
    sections = defaultdict(list)
    for item in items:
        sections[item.section].append(item)

    for sec_name, sec_items in sections.items():
        totals = [it for it in sec_items if it.is_total and not it.is_grand_total]
        grand_totals = [it for it in sec_items if it.is_grand_total]
        data = [it for it in sec_items
                if not it.is_total and not it.is_section_header]

        for t in totals:
            if t.total is None:
                continue
            computed = sum(d.total for d in data if d.total is not None)
            if computed and abs(computed - t.total) > 1:
                checks.append({
                    'section': sec_name,
                    'computed': computed,
                    'stated': t.total,
                    'diff': computed - t.total,
                    'type': 'section',
                })

        for t in grand_totals:
            if t.total is None:
                continue
            computed = sum(d.total for d in data if d.total is not None)
            if computed and abs(computed - t.total) > 1:
                checks.append({
                    'section': sec_name,
                    'computed': computed,
                    'stated': t.total,
                    'diff': computed - t.total,
                    'type': 'grand_total',
                })

    return checks


# ═══════════════════════════════════════════════════════════════════════════════
#  Export
# ═══════════════════════════════════════════════════════════════════════════════

def _style_cell(cell, is_header=False, is_total=False, alt=False):
    thin = Border(left=Side('thin'), right=Side('thin'),
                  top=Side('thin'), bottom=Side('thin'))
    cell.border = thin
    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    if is_header:
        cell.fill = PatternFill('solid', fgColor='D9E1F2')
        cell.font = Font(bold=True, size=10)
    elif is_total:
        cell.fill = PatternFill('solid', fgColor='E2EFDA')
        cell.font = Font(bold=True, size=10)
    elif alt:
        cell.fill = PatternFill('solid', fgColor='F5F5F5')


def export_excel(pages: list[dict], output_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    for pd_ in pages:
        ptype = pd_['type']
        pnum = pd_['page_num'] + 1
        if ptype == 'empty':
            continue

        ws = wb.create_sheet(title=f'Page {pnum}')

        if ptype in ('text', 'cover'):
            ws.cell(row=1, column=1, value=pd_.get('raw_text', ''))
            continue

        if ptype == 'table':
            for ti, table in enumerate(pd_.get('tables', [])):
                base_row = 1
                if ti > 0:
                    base_row = len(table.rows) + 3

                for ci, h in enumerate(table.header):
                    cell = ws.cell(row=base_row, column=ci + 1,
                                   value=h.replace('\n', ' '))
                    _style_cell(cell, is_header=True)
                base_row += 1

                for ri, row in enumerate(table.rows):
                    is_total = is_total_text(' '.join(row).strip())
                    for ci, val in enumerate(row):
                        cell = ws.cell(row=base_row + ri, column=ci + 1,
                                       value=val.replace('\n', ' '))
                        _style_cell(cell, is_total=is_total,
                                    alt=ri % 2 == 1)

                items = table.__dict__.get('budget_items', [])
                if items:
                    br = base_row + len(table.rows) + 2
                    ws.cell(row=br, column=1,
                            value='Flattened Budget Items').font = Font(
                        bold=True, color='003366')
                    br += 1
                    cols = ['Page', 'Section', 'Code', 'Description',
                            'Year Actual', 'Year Revised', 'Year Estimate',
                            'Total', 'Current Exp', 'Capital Exp',
                            'Financial', 'Is Total', 'Is Section',
                            'Is Grand Total']
                    for ci, c in enumerate(cols):
                        cell = ws.cell(row=br, column=ci + 1, value=c)
                        cell.fill = PatternFill('solid', fgColor='BDD7EE')
                        cell.font = Font(bold=True, size=9)
                    br += 1
                    for item in items:
                        vals = [item.page, item.section, item.code,
                                item.description,
                                item.year_actual, item.year_revised,
                                item.year_estimate,
                                item.total, item.current_exp,
                                item.capital_exp, item.financial,
                                'Yes' if item.is_total else '',
                                'Yes' if item.is_section_header else '',
                                'Yes' if item.is_grand_total else '']
                        for ci, v in enumerate(vals):
                            ws.cell(row=br, column=ci + 1, value=v)
                        br += 1

                checks = table.__dict__.get('cross_checks', [])
                if checks:
                    cr = br + 2
                    ws.cell(row=cr, column=1,
                            value='Cross-Verification').font = Font(
                        bold=True, color='CC0000')
                    cr += 1
                    for i, h in enumerate(['Type', 'Section', 'Computed',
                                           'Stated', 'Diff']):
                        cell = ws.cell(row=cr, column=i + 1, value=h)
                        cell.fill = PatternFill('solid', fgColor='FF9999')
                    cr += 1
                    for ch in checks:
                        ws.cell(row=cr, column=1, value=ch.get('type', ''))
                        ws.cell(row=cr, column=2, value=ch['section'])
                        ws.cell(row=cr, column=3,
                                value=round(ch['computed'], 2))
                        ws.cell(row=cr, column=4,
                                value=round(ch['stated'], 2))
                        ws.cell(row=cr, column=5,
                                value=round(ch['diff'], 2))
                        cr += 1

    if not wb.worksheets:
        ws = wb.create_sheet('Output')
        ws.cell(row=1, column=1, value='No data found in PDF.')

    wb.save(output_path)
    print(f'[excel] Written to {output_path}', file=sys.stderr)


def export_csv(pages: list[dict], output_dir: str, stem: str):
    """Export flattened budget items to CSV, one per page,
    with section column populated for every row."""
    for pd_ in pages:
        ptype = pd_['type']
        pnum = pd_['page_num'] + 1
        if ptype == 'empty':
            continue

        def _val(s):
            s = str(s).replace('"', '""')
            if ',' in s or '"' in s or '\n' in s:
                return f'"{s}"'
            return s

        path = Path(output_dir) / f'{stem}-page{pnum}.csv'

        if ptype in ('text', 'cover'):
            safe = str(pd_.get('raw_text', '')).replace('"', '""')
            path.write_text(f'"{safe}"', encoding='utf-8')

        elif ptype == 'table':
            items = []
            for table in pd_.get('tables', []):
                items.extend(table.__dict__.get('budget_items', []))
            if items:
                lines = ['Section,Code,Description,Year Actual,Year Revised,'
                         'Year Estimate,Total,Current Exp,Capital Exp,'
                         'Financial,Is Total,Is Grand Total']
                for item in items:
                    d = item.description.replace('"', '""')
                    lines.append(
                        f'{_val(item.section)},{_val(item.code)},'
                        f'"{d}",'
                        f'{item.year_actual or ""},{item.year_revised or ""},'
                        f'{item.year_estimate or ""},{item.total or ""},'
                        f'{item.current_exp or ""},{item.capital_exp or ""},'
                        f'{item.financial or ""},'
                        f'{"Yes" if item.is_total else ""},'
                        f'{"Yes" if item.is_grand_total else ""}'
                    )
            else:
                lines = ['(no budget items found)']
            path.write_text('\n'.join(lines), encoding='utf-8')

        print(f'[csv]   {path.name}', file=sys.stderr)


def export_sqlite(pages: list[dict], output_path: str):
    """Export all pages to a unified SQLite database with a flattened
    staging schema for clean queries and CSV export."""
    conn = sqlite3.connect(str(output_path))
    conn.execute('PRAGMA journal_mode=WAL')
    conn.execute('PRAGMA foreign_keys=ON')

    # Drop legacy tables from previous schema versions
    for tbl in ('pages', 'tables', 'rows'):
        conn.execute(f'DROP TABLE IF EXISTS {tbl}')

    conn.execute("""
        CREATE TABLE IF NOT EXISTS raw_budget_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            page_number INTEGER NOT NULL,
            table_type TEXT,
            inferred_section TEXT,
            budget_code TEXT,
            description TEXT,
            year_actual REAL,
            year_revised REAL,
            year_estimate REAL,
            total REAL,
            current_exp REAL,
            capital_exp REAL,
            financial REAL,
            is_total_row INTEGER DEFAULT 0,
            is_grand_total_row INTEGER DEFAULT 0,
            is_section_header INTEGER DEFAULT 0,
            scale_multiplier INTEGER DEFAULT 1
        )
    """)

    for pd_ in pages:
        ptype = pd_['type']
        pnum = pd_['page_num'] + 1
        if ptype != 'table':
            continue
        for table in pd_.get('tables', []):
            items = table.__dict__.get('budget_items', [])
            table_type = 'detail' if any(item.code for item in items) else 'summary'
            for item in items:
                scale = 1
                conn.execute("""
                    INSERT INTO raw_budget_lines
                        (page_number, table_type, inferred_section,
                         budget_code, description,
                         year_actual, year_revised, year_estimate,
                         total, current_exp, capital_exp, financial,
                         is_total_row, is_grand_total_row,
                         is_section_header, scale_multiplier)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?,
                            ?, ?, ?, ?)
                """, (
                    pnum, table_type, item.section,
                    item.code, item.description,
                    item.year_actual, item.year_revised,
                    item.year_estimate,
                    item.total, item.current_exp,
                    item.capital_exp, item.financial,
                    1 if item.is_total else 0,
                    1 if item.is_grand_total else 0,
                    1 if item.is_section_header else 0,
                    scale,
                ))

    conn.commit()
    conn.close()
    print(f'[sqlite] Written to {output_path}', file=sys.stderr)


# ═══════════════════════════════════════════════════════════════════════════════
#  Main Pipeline
# ═══════════════════════════════════════════════════════════════════════════════

def _extract_raw_cids(pdf_path: str, page_num: int) -> list[int]:
    from pypdf import PdfReader
    reader = PdfReader(pdf_path)
    if page_num >= len(reader.pages):
        return []
    page = reader.pages[page_num]
    content = page.get_contents()
    if content is None:
        return []
    try:
        data = content.get_data()
    except Exception:
        data = None
    if not data:
        return []
    if isinstance(data, bytes):
        data = data.decode('latin-1')

    cid_seq = []
    for h in re.findall(r'<([0-9A-Fa-f]+)>', data):
        if len(h) <= 4:
            try:
                c = int(h, 16)
                if 0 < c < 1000:
                    cid_seq.append(c)
            except ValueError:
                pass
        else:
            for i in range(0, len(h), 4):
                try:
                    c = int(h[i:i+4], 16)
                    if 0 < c < 1000:
                        cid_seq.append(c)
                except ValueError:
                    pass
    return cid_seq


def fix_ufffd_chars(pdf_path: str, page, page_num: int,
                    font_mapper) -> dict:
    fixes = {}
    cid_seq = _extract_raw_cids(pdf_path, page_num)
    if not cid_seq:
        return fixes

    chars_sorted = sorted(page.chars, key=lambda c: (c['top'], c['x0']))
    cid_idx = 0
    for pc in chars_sorted:
        if cid_idx >= len(cid_seq):
            break
        if pc.get('text') == '\ufffd':
            cid = cid_seq[cid_idx]
            decoded = font_mapper.cid_to_uni.get(cid, '')
            if decoded:
                fixes[(round(pc['x0'], 1), round(pc['top'], 1))] = decoded
        cid_idx += 1

    return fixes


def decode_cell_text(cell_text: str, font_mapper: FontMapper,
                     uffd_fixes: dict = None) -> str:
    decoded = font_mapper.decode_text(cell_text or '')
    if '\ufffd' not in decoded or not uffd_fixes:
        return decoded
    return decoded.replace('\ufffd', '')


def extract_pdf(pdf_path: str, font_mapper: FontMapper,
                max_pages: int | None = None) -> list[dict]:
    pages: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for pi, page in enumerate(pdf.pages):
            if max_pages and pi >= max_pages:
                break

            result = {'page_num': pi, 'type': 'empty'}

            uffd_fixes = fix_ufffd_chars(pdf_path, page, pi, font_mapper)

            raw_text = page.extract_text() or ''
            if uffd_fixes:
                sorted_chars = sorted(page.chars,
                                      key=lambda c: (c['top'], c['x0']))
                fixed_parts = []
                for c in sorted_chars:
                    txt = c.get('text', '')
                    if txt == '\ufffd':
                        key = (round(c['x0'], 1), round(c['top'], 1))
                        fixed_parts.append(uffd_fixes.get(key, ''))
                    else:
                        fixed_parts.append(txt)
                fixed_raw = ''.join(fixed_parts)
                result['raw_text'] = font_mapper.decode_text(fixed_raw)
            else:
                result['raw_text'] = font_mapper.decode_text(raw_text)

            # TOC gatekeeper: skip non-budget pages
            if not is_valid_budget_page(result['raw_text']):
                result['type'] = 'cover'
                pages.append(result)
                continue

            tables = page.find_tables()
            decoded_tables = []
            for t in tables:
                data = t.extract()
                if not data or len(data) < 2:
                    continue
                header = [decode_cell_text(c, font_mapper, uffd_fixes)
                          for c in data[0]]
                rows = []
                for row in data[1:]:
                    rows.append([decode_cell_text(c, font_mapper, uffd_fixes)
                                 for c in row])
                first_h = (header[0] or '').strip()
                if first_h and re.match(r'^\d+', first_h):
                    rows.insert(0, header)
                    header = ['कोड', 'विवरण', 'यथार्थ', 'संशोधित',
                              'अनुमान', 'जम्मा', 'अन्य']
                decoded_tables.append(PageTable(
                    page_num=pi,
                    bbox=t.bbox,
                    header=header,
                    rows=rows,
                ))

            if decoded_tables:
                result['type'] = 'table'
                result['tables'] = decoded_tables
            elif result['raw_text'].strip():
                cover = len([l for l in result['raw_text'].split('\n')
                             if l.strip()]) <= 5
                result['type'] = 'cover' if cover else 'text'
            else:
                result['type'] = 'empty'

            pages.append(result)

    return pages


def build_font_mapper(pdf_path: str) -> FontMapper:
    from pypdf import PdfReader
    mapper = FontMapper()
    reader = PdfReader(pdf_path)
    seen = set()
    for page in reader.pages:
        fonts = page['/Resources'].get('/Font',
                                       {}) if '/Resources' in page else {}
        for font_obj in fonts.values():
            if '/ToUnicode' in font_obj:
                base = font_obj.get('/BaseFont', '')
                if base in seen:
                    continue
                seen.add(base)
                tu = font_obj['/ToUnicode']
                data = tu.get_data()
                if isinstance(data, bytes):
                    data = data.decode('latin-1')
                mapper.ingest_cmap(str(data))
    n_cid = len(mapper.cid_to_uni)
    print(f'[init] Font mapper: {len(seen)} CMap(s), {n_cid} CID mappings',
          file=sys.stderr)
    return mapper


def run_pipeline(pdf_path: str, max_pages: int | None = None) -> list[dict]:
    mapper = build_font_mapper(pdf_path)
    pages = extract_pdf(pdf_path, mapper, max_pages)

    prev_section = ''

    for pd_ in pages:
        ptype = pd_['type']
        pnum = pd_['page_num'] + 1
        if ptype == 'table':
            all_items = []
            last_section = prev_section

            # Detect scale from page raw text
            scale = detect_scale_multiplier(pd_.get('raw_text', ''))

            for table in pd_['tables']:
                items = classify_budget_table(table, last_section, scale)
                for item in items:
                    if item.section:
                        last_section = item.section
                table.__dict__['budget_items'] = items
                all_items.extend(items)
                checks = cross_verify(items)
                table.__dict__['cross_checks'] = checks
                if checks:
                    for ch in checks:
                        s = ch['section']
                        c = ch['computed']
                        st = ch['stated']
                        d = ch['diff']
                        t = ch.get('type', 'section')
                        print(f'[verify] P{pnum} "{s}" ({t}): '
                              f'computed={c:.0f} stated={st:.0f} '
                              f'diff={d:.0f}',
                              file=sys.stderr)
            pd_['items'] = all_items
            if all_items:
                for item in reversed(all_items):
                    s = item.section
                    if s and not item.is_total and len(s) >= 2 \
                       and re.match(r'[\u0900-\u097F]', s):
                        prev_section = s
                        break
        print(f'[page]  P{pnum}: {ptype}', file=sys.stderr)

    return pages


def main():
    parser = argparse.ArgumentParser(
        description='Nepali budget PDF → XLSX/CSV/SQLite '
                    '(vector-layer extraction)')
    parser.add_argument('pdf', help='Input PDF')
    parser.add_argument('--max-pages', type=int, default=None,
                        help='Only first N pages')
    parser.add_argument('--output', '-o', help='Output path')
    group = parser.add_mutually_exclusive_group()
    group.add_argument('--xlsx', action='store_true', help='Excel (default)')
    group.add_argument('--csv', action='store_true',
                       help='CSV per page (flat budget items)')
    group.add_argument('--sqlite', action='store_true', help='SQLite DB')
    args = parser.parse_args()

    pdf = Path(args.pdf)
    if not pdf.exists():
        print(f'Error: {pdf} not found', file=sys.stderr)
        sys.exit(1)

    Path('output').mkdir(exist_ok=True)
    stem = pdf.stem
    pages = run_pipeline(str(pdf), args.max_pages)

    if args.csv:
        export_csv(pages, 'output', stem)
    elif args.sqlite:
        export_sqlite(pages, args.output or f'output/{stem}.db')
    else:
        export_excel(pages, args.output or f'output/{stem}_v2.xlsx')

    print(f'[done] {len(pages)} page(s) processed.', file=sys.stderr)


if __name__ == '__main__':
    main()
